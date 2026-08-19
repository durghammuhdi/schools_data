import csv
import os
import flet as ft


FILENAME = "schools_data.csv"
EXPORT_FILENAME = "تقرير_الإشراف_التربوي.csv"
HEADERS = [
    "المدرسة",
    "الاختصاص",
    "الملاك المطلوب",
    "العدد الحالي",
    "الحالة",
    "الفرق",
    "القضاء",
    "المرحلة الدراسية",
    "الملاحظات",
    "أسماء مدرّسي المادة",
]


def main(page: ft.Page):
    page.title = "نظام الإشراف التربوي - إدارات المدارس"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.padding = 0
    page.rtl = True

    editing_index = [-1]

    # إنشاء ملف البيانات إذا لم يكن موجوداً
    if not os.path.exists(FILENAME):
        with open(FILENAME, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(HEADERS)

    def load_records():
        records = []

        if os.path.exists(FILENAME):
            with open(FILENAME, mode="r", encoding="utf-8-sig") as file:
                reader = csv.reader(file)
                next(reader, None)

                for row in reader:
                    if row and len(row) >= 6:
                        row = row + [""] * (10 - len(row))
                        records.append(row)

        return records

    def save_all_records(records):
        with open(FILENAME, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(HEADERS)
            writer.writerows(records)

    # بطاقات الإحصائيات
    card_total = ft.Text(
        "0",
        size=22,
        weight=ft.FontWeight.BOLD,
        color="blue",
    )

    card_vacant = ft.Text(
        "0",
        size=22,
        weight=ft.FontWeight.BOLD,
        color="red",
    )

    card_surplus = ft.Text(
        "0",
        size=22,
        weight=ft.FontWeight.BOLD,
        color="orange",
    )

    status_label = ft.Text(
        value="",
        color="green",
        size=14,
    )

    # حقول الإدخال
    txt_school = ft.TextField(
        label="اسم المدرسة",
        width=240,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_spec = ft.TextField(
        label="الاختصاص",
        width=220,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_district = ft.TextField(
        label="القضاء",
        width=180,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_stage = ft.TextField(
        label="المرحلة الدراسية",
        width=180,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_notes = ft.TextField(
        label="الملاحظات",
        width=240,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_teachers = ft.TextField(
        label="أسماء مدرّسي المادة",
        width=280,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_req = ft.TextField(
        label="الملاك المطلوب",
        width=130,
        keyboard_type=ft.KeyboardType.NUMBER,
        text_align=ft.TextAlign.RIGHT,
    )

    txt_curr = ft.TextField(
        label="العدد الحالي",
        width=130,
        keyboard_type=ft.KeyboardType.NUMBER,
        text_align=ft.TextAlign.RIGHT,
    )

    # =========================================================
    # البحث والتصفية
    # =========================================================

    txt_search_school = ft.TextField(
        label="تصفية حسب المدرسة...",
        width=220,
        prefix_icon=ft.Icons.SEARCH,
        text_align=ft.TextAlign.RIGHT,
        on_change=lambda e: refresh_table(),
    )

    # يتم تحديث الخيارات تلقائياً من الاختصاصات الموجودة في ملف البيانات
    filter_dropdown = ft.Dropdown(
        label="تصفية حسب الاختصاص",
        width=220,
        value="الكل",
        options=[ft.dropdown.Option("الكل")],
        on_select=lambda e: refresh_table(),
    )

    # فلتر الحالة
    status_filter = ft.Dropdown(
        label="تصفية حسب الحالة",
        width=190,
        value="الكل",
        options=[
            ft.dropdown.Option("الكل"),
            ft.dropdown.Option("شاغر"),
            ft.dropdown.Option("فيض"),
            ft.dropdown.Option("مكتمل"),
        ],
        on_select=lambda e: refresh_table(),
    )

    def clear_filters(e=None):
        """مسح جميع الفلاتر وإظهار كل السجلات."""
        txt_search_school.value = ""
        filter_dropdown.value = "الكل"
        status_filter.value = "الكل"
        refresh_table()

    btn_clear_filters = ft.Button(
        content="مسح الفلاتر",
        icon=ft.Icons.CLEAR_ALL,
        on_click=clear_filters,
    )

    def normalize_text(value):
        """توحيد النص حتى لا يفشل الفلتر بسبب المسافات الزائدة."""
        if value is None:
            return ""
        return " ".join(str(value).strip().split()).lower()

    def update_filter_options():
        """
        إنشاء قائمة الاختصاصات تلقائياً من السجلات.
        أي اختصاص جديد تتم إضافته يظهر في عامل التصفية
        بدون الحاجة إلى تعديل الكود.
        """
        current_value = filter_dropdown.value or "الكل"

        records = load_records()

        specialties = sorted(
            {
                " ".join(str(row[1]).strip().split())
                for row in records
                if len(row) >= 2 and str(row[1]).strip()
            }
        )

        options = [ft.dropdown.Option("الكل")]
        options.extend(ft.dropdown.Option(spec) for spec in specialties)

        filter_dropdown.options = options

        # إذا كان الاختصاص القديم لم يعد موجوداً، نرجع إلى الكل
        if current_value == "الكل" or current_value in specialties:
            filter_dropdown.value = current_value
        else:
            filter_dropdown.value = "الكل"

    # جدول البيانات
    data_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("إجراءات", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("الفرق", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("الحالة", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("الحالي", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("الملاك", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("الاختصاص", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("القضاء", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("المرحلة", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("الملاحظات", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("مدرّسو المادة", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("المدرسة", weight=ft.FontWeight.BOLD)),
        ],
        rows=[],
    )

    def update_dashboard():
        records = load_records()

        total_schools = len(
            set(r[0].strip() for r in records if r[0].strip())
        )

        vacant_count = sum(
            1 for r in records if r[4].strip() == "شاغر"
        )

        surplus_count = sum(
            1 for r in records if r[4].strip() == "فيض"
        )

        card_total.value = str(total_schools)
        card_vacant.value = str(vacant_count)
        card_surplus.value = str(surplus_count)

    def refresh_table(e=None):
        data_table.rows.clear()
        records = load_records()

        # تحديث الاختصاصات الموجودة فعلياً في البيانات
        update_filter_options()

        query_school = normalize_text(
            txt_search_school.value
            if txt_search_school.value
            else ""
        )

        selected_spec = normalize_text(
            filter_dropdown.value
            if filter_dropdown.value and filter_dropdown.value != "الكل"
            else ""
        )

        selected_status = normalize_text(
            status_filter.value
            if status_filter.value and status_filter.value != "الكل"
            else ""
        )

        for i, row in enumerate(records):
            idx = i
            r_data = row

            school_name = normalize_text(row[0])
            spec_name = normalize_text(row[1])

            # البحث باسم المدرسة: يسمح بكتابة جزء من الاسم
            if query_school and query_school not in school_name:
                continue

            # التصفية حسب الاختصاص
            if selected_spec and selected_spec != spec_name:
                continue

            # التصفية حسب الحالة
            row_status = normalize_text(row[4])
            if selected_status and selected_status != row_status:
                continue

            data_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(
                            ft.Row(
                                controls=[
                                    ft.IconButton(
                                        icon=ft.Icons.EDIT,
                                        icon_color="blue",
                                        tooltip="تعديل",
                                        on_click=lambda e, idx=idx, r=r_data:
                                        start_edit(idx, r),
                                    ),
                                    ft.IconButton(
                                        icon=ft.Icons.DELETE,
                                        icon_color="red",
                                        tooltip="حذف",
                                        on_click=lambda e, idx=idx:
                                        delete_record(idx),
                                    ),
                                ],
                                alignment=ft.MainAxisAlignment.END,
                            )
                        ),
                        ft.DataCell(ft.Text(row[5])),
                        ft.DataCell(ft.Text(row[4])),
                        ft.DataCell(ft.Text(row[3])),
                        ft.DataCell(ft.Text(row[2])),
                        ft.DataCell(ft.Text(row[1])),
                        ft.DataCell(ft.Text(row[6])),
                        ft.DataCell(ft.Text(row[7])),
                        ft.DataCell(ft.Text(row[8])),
                        ft.DataCell(ft.Text(row[9])),
                        ft.DataCell(ft.Text(row[0])),
                    ]
                )
            )

        update_dashboard()
        page.update()

    txt_search_school.on_change = refresh_table
    filter_dropdown.on_select = refresh_table

    def start_edit(index, row_data):
        editing_index[0] = index

        txt_school.value = row_data[0]
        txt_spec.value = row_data[1]
        txt_req.value = row_data[2]
        txt_curr.value = row_data[3]
        txt_district.value = row_data[6]
        txt_stage.value = row_data[7]
        txt_notes.value = row_data[8]
        txt_teachers.value = row_data[9]

        btn_add.content = "تحديث السجل"
        btn_cancel.visible = True

        status_label.value = f"جاري تعديل: {row_data[0]}"
        status_label.color = "blue"

        page.update()

    def cancel_edit(e=None):
        editing_index[0] = -1

        txt_school.value = ""
        txt_spec.value = ""
        txt_req.value = ""
        txt_curr.value = ""
        txt_district.value = ""
        txt_stage.value = ""
        txt_notes.value = ""
        txt_teachers.value = ""

        btn_add.content = "إضافة سجل"
        btn_cancel.visible = False

        status_label.value = ""
        page.update()

    def delete_record(index):
        records = load_records()

        if 0 <= index < len(records):
            records.pop(index)
            save_all_records(records)

            refresh_table()

            status_label.value = "تم حذف السجل بنجاح!"
            status_label.color = "orange"
            page.update()

    def save_or_update(e):
        if (
            not txt_school.value
            or not txt_spec.value
            or not txt_req.value
            or not txt_curr.value
        ):
            status_label.value = "يرجى ملء جميع الحقول!"
            status_label.color = "red"
            page.update()
            return

        try:
            req = int(txt_req.value)
            curr = int(txt_curr.value)
        except ValueError:
            status_label.value = "الملاك والأعداد يجب أن تكون أرقاماً فقط!"
            status_label.color = "red"
            page.update()
            return

        diff = curr - req

        if diff < 0:
            status = "شاغر"
        elif diff > 0:
            status = "فيض"
        else:
            status = "مكتمل"

        new_row = [
            txt_school.value.strip(),
            txt_spec.value.strip(),
            str(req),
            str(curr),
            status,
            str(abs(diff)),
            txt_district.value.strip(),
            txt_stage.value.strip(),
            txt_notes.value.strip(),
            txt_teachers.value.strip(),
        ]

        records = load_records()

        if editing_index[0] != -1:
            if editing_index[0] < len(records):
                records[editing_index[0]] = new_row

            save_all_records(records)

            status_label.value = "تم تحديث السجل بنجاح!"
            status_label.color = "green"
        else:
            records.append(new_row)
            save_all_records(records)

            status_label.value = "تمت إضافة السجل بنجاح!"
            status_label.color = "green"

        cancel_edit()
        refresh_table()

    # ---------------------------------------------------------
    # حل مشكلة FilePicker:
    # في إصدارات Flet الحديثة أصبح FilePicker خدمة async،
    # لذلك لا نضيفه إلى page.overlay ولا نستخدم on_result.
    # ---------------------------------------------------------

    file_picker = ft.FilePicker()

    async def export_data_to_file(e):
        records = load_records()

        # إنشاء محتوى CSV في الذاكرة
        import io

        output = io.StringIO(newline="")
        writer = csv.writer(output)

        writer.writerow(HEADERS)

        writer.writerows(records)

        csv_bytes = output.getvalue().encode("utf-8-sig")

        try:
            saved_path = await file_picker.save_file(
                dialog_title="حفظ تقرير الإشراف التربوي",
                file_name=EXPORT_FILENAME,
                file_type=ft.FilePickerFileType.CUSTOM,
                allowed_extensions=["csv"],
                src_bytes=csv_bytes,
            )

            if saved_path:
                status_label.value = (
                    f"تم تصدير التقرير بنجاح إلى: {saved_path}"
                )
                status_label.color = "green"
            else:
                status_label.value = "تم إلغاء عملية التصدير."
                status_label.color = "orange"

        except Exception as ex:
            status_label.value = f"حدث خطأ أثناء التصدير: {str(ex)}"
            status_label.color = "red"

        page.update()

    async def export_data_to_excel(e):
        records = load_records()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            status_label.value = (
                "لتصدير Excel ثبّت الحزمة أولاً: pip install openpyxl"
            )
            status_label.color = "red"
            page.update()
            return

        try:
            saved_path = await file_picker.save_file(
                dialog_title="حفظ تقرير Excel",
                file_name="تقرير_الإشراف_التربوي.xlsx",
                file_type=ft.FilePickerFileType.CUSTOM,
                allowed_extensions=["xlsx"],
            )

            if not saved_path:
                status_label.value = "تم إلغاء عملية التصدير."
                status_label.color = "orange"
                page.update()
                return

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "تقرير الإشراف"

            worksheet.append(HEADERS)

            for row in records:
                worksheet.append(row)

            header_fill = PatternFill(
                fill_type="solid",
                fgColor="2563EB",
            )
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            for column_cells in worksheet.columns:
                column_letter = column_cells[0].column_letter
                max_length = max(
                    len(str(cell.value or "")) for cell in column_cells
                )
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 12),
                    35,
                )

            workbook.save(saved_path)

            status_label.value = f"تم تصدير تقرير Excel بنجاح إلى: {saved_path}"
            status_label.color = "green"

        except Exception as ex:
            status_label.value = f"حدث خطأ أثناء تصدير Excel: {str(ex)}"
            status_label.color = "red"

        page.update()

    # الأزرار
    btn_add = ft.Button(
        content="إضافة سجل",
        on_click=save_or_update,
        icon=ft.Icons.ADD,
    )

    btn_cancel = ft.Button(
        content="إلغاء التعديل",
        on_click=cancel_edit,
        icon=ft.Icons.CLOSE,
        visible=False,
    )

    btn_export = ft.Button(
        content="تصدير التقرير",
        on_click=export_data_to_file,
        icon=ft.Icons.SAVE,
    )

    btn_export_excel = ft.Button(
        content="تصدير Excel",
        on_click=export_data_to_excel,
        icon=ft.Icons.SAVE,
    )

    def show_status_records(status, title, color):
        records = [
            row for row in load_records()
            if len(row) >= 6 and row[4].strip() == status
        ]

        if records:
            record_controls = [
                ft.Container(
                    padding=8,
                    border_radius=6,
                    content=ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Text(
                                f"الفرق: {row[5]}",
                                color=color,
                                weight=ft.FontWeight.BOLD,
                            ),
                            ft.Column(
                                spacing=2,
                                horizontal_alignment=ft.CrossAxisAlignment.END,
                                controls=[
                                    ft.Text(
                                        row[0],
                                        weight=ft.FontWeight.BOLD,
                                    ),
                                    ft.Text(
                                        f"الاختصاص: {row[1]}",
                                        color="grey",
                                        size=12,
                                    ),
                                ],
                            ),
                        ],
                    ),
                )
                for row in records
            ]
        else:
            record_controls = [
                ft.Container(
                    padding=20,
                    content=ft.Text(
                        f"لا توجد مدارس ضمن {status} حالياً.",
                        color="grey",
                        size=15,
                    ),
                )
            ]

        def close_dialog(e):
            dialog.open = False
            page.update()

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text(title, color=color, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=520,
                height=350,
                content=ft.Column(
                    scroll=ft.ScrollMode.AUTO,
                    spacing=8,
                    controls=record_controls,
                ),
            ),
            actions=[
                ft.Button(
                    content="إغلاق",
                    on_click=close_dialog,
                )
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        page.show_dialog(dialog)

    def make_stat_card(
        title,
        value_widget,
        icon_name,
        icon_color,
        on_click=None,
    ):
        card = ft.Card(
            elevation=2,
            content=ft.Container(
                padding=15,
                width=200,
                content=ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    controls=[
                        ft.Icon(
                            icon_name,
                            color=icon_color,
                            size=30,
                        ),
                        ft.Column(
                            horizontal_alignment=ft.CrossAxisAlignment.END,
                            controls=[
                                ft.Text(
                                    title,
                                    size=12,
                                    color="gray",
                                ),
                                value_widget,
                            ],
                        ),
                    ],
                ),
            ),
        )

        if on_click:
            return ft.GestureDetector(
                on_tap=on_click,
                content=card,
            )

        return card

    stats_row = ft.Row(
        alignment=ft.MainAxisAlignment.END,
        wrap=True,
        controls=[
            make_stat_card(
                "حالات الفيض",
                card_surplus,
                ft.Icons.ADD_ALERT,
                "orange",
                on_click=lambda e: show_status_records(
                    "فيض",
                    "المدارس التي فيها فيض",
                    "orange",
                ),
            ),
            make_stat_card(
                "حالات الشاغر",
                card_vacant,
                ft.Icons.WARNING,
                "red",
                on_click=lambda e: show_status_records(
                    "شاغر",
                    "المدارس التي فيها شاغر",
                    "red",
                ),
            ),
            make_stat_card(
                "إجمالي المدارس",
                card_total,
                ft.Icons.SCHOOL,
                "blue",
            ),
        ],
    )

    footer_section = ft.Row(
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
        controls=[
            ft.Text(
                "برمجة وتطوير: ضرغام مهدي صلال",
                size=14,
                color="#334155",
                weight=ft.FontWeight.BOLD,
            ),
            ft.Text(
                "نظام الإشراف التربوي © 2026",
                size=12,
                color="grey",
            ),
        ],
    )

    main_layout = ft.Container(
        expand=True,
        gradient=ft.LinearGradient(
            begin=ft.Alignment(-1, -1),
            end=ft.Alignment(1, 1),
            colors=["#eef2f3", "#8e9eab"],
        ),
        padding=25,
        content=ft.Column(
            scroll=ft.ScrollMode.AUTO,
            alignment=ft.MainAxisAlignment.START,
            horizontal_alignment=ft.CrossAxisAlignment.END,
            controls=[
                ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    controls=[
                         ft.Row(
                             controls=[
                                 btn_export_excel,
                                 btn_export,
                             ],
                             spacing=8,
                         ),
                        ft.Text(
                            "نظام الإشراف التربوي - إدارات المدارس",
                            size=24,
                            weight=ft.FontWeight.BOLD,
                            color="#1e293b",
                        ),
                    ],
                ),

                ft.Container(
                    alignment=ft.Alignment(1, 0),
                    padding=ft.Padding(0, 4, 0, 4),
                    content=ft.Text(
                        "برمجة وتطوير: ضرغام مهدي صلال",
                        size=15,
                        weight=ft.FontWeight.BOLD,
                        color="#2563eb",
                    ),
                ),

                ft.Divider(
                    height=15,
                    color="transparent",
                ),

                stats_row,

                ft.Divider(
                    height=15,
                    color="transparent",
                ),

                ft.Card(
                    elevation=3,
                    content=ft.Container(
                        padding=20,
                        border_radius=10,
                        content=ft.Column(
                            horizontal_alignment=ft.CrossAxisAlignment.END,
                            controls=[
                                ft.Row(
                                    controls=[
                                        txt_spec,
                                        txt_school,
                                    ],
                                    alignment=ft.MainAxisAlignment.END,
                                    wrap=True,
                                ),

                                ft.Row(
                                    controls=[
                                        txt_notes,
                                        txt_stage,
                                        txt_district,
                                    ],
                                    alignment=ft.MainAxisAlignment.END,
                                    wrap=True,
                                ),

                                ft.Row(
                                    controls=[txt_teachers],
                                    alignment=ft.MainAxisAlignment.END,
                                    wrap=True,
                                ),

                                ft.Row(
                                    controls=[
                                        btn_cancel,
                                        btn_add,
                                        txt_curr,
                                        txt_req,
                                    ],
                                    alignment=ft.MainAxisAlignment.END,
                                    wrap=True,
                                ),

                                status_label,
                            ],
                        ),
                    ),
                ),

                ft.Divider(
                    height=15,
                    color="transparent",
                ),

                ft.Card(
                    elevation=3,
                    content=ft.Container(
                        padding=20,
                        border_radius=10,
                        content=ft.Column(
                            horizontal_alignment=ft.CrossAxisAlignment.END,
                            controls=[
                                ft.Row(
                                    controls=[
                                        ft.Row(
                                            controls=[
                                                btn_clear_filters,
                                                status_filter,
                                                filter_dropdown,
                                                txt_search_school,
                                            ],
                                            spacing=8,
                                            wrap=True,
                                        ),
                                        ft.Text(
                                            "جدول المدارس المحفوظة",
                                            size=18,
                                            weight=ft.FontWeight.BOLD,
                                            color="#334155",
                                        ),
                                    ],
                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                                ),

                                ft.Row(
                                    controls=[data_table],
                                    alignment=ft.MainAxisAlignment.END,
                                    scroll=ft.ScrollMode.ALWAYS,
                                ),
                            ],
                        ),
                    ),
                ),

                ft.Divider(
                    height=15,
                    color="transparent",
                ),

                footer_section,
            ],
        ),
    )

    page.add(main_layout)
    refresh_table()


if __name__ == "__main__":
    ft.run(main)
